# your_package/document_processor/excel_handler.py
"""
Excel Handler - Excel Document Processor (XLSX/XLS)

Main Features:
- Metadata extraction (title, author, subject, keywords, creation date, modification date, etc.)
- Text extraction (direct parsing via openpyxl/xlrd)
- Table extraction (Markdown or HTML conversion based on merged cells)
- Inline image extraction and local storage
- Chart processing (1st priority: convert to table, 2nd priority: matplotlib image)
- Multi-sheet support

Class-based Handler:
- ExcelHandler class inherits from BaseHandler to manage config/image_processor
"""
from __future__ import annotations

import logging
import os
from typing import TYPE_CHECKING, Any, Dict, List, Optional, Set

from libs.core.processor.base_handler import BaseHandler
from libs.core.functions.img_processor import ImageProcessor

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from libs.core.document_processor import CurrentFile
from libs.core.processor.excel_helper import (
    # Textbox
    extract_textboxes_from_xlsx,
    # Metadata
    extract_xlsx_metadata,
    extract_xls_metadata,
    format_metadata,
    # Chart
    extract_charts_from_xlsx,
    process_chart,
    extract_chart_info_basic,
    # Image
    extract_images_from_xlsx,
    get_sheet_images,
    # Table
    convert_xlsx_sheet_to_table,
    convert_xls_sheet_to_table,
    # Object Detection (개별 테이블 청킹)
    convert_xlsx_objects_to_tables,
    convert_xls_objects_to_tables,
)

import xlrd
from openpyxl import load_workbook

logger = logging.getLogger("document-processor")


# ============================================================================
# ExcelHandler Class
# ============================================================================

class ExcelHandler(BaseHandler):
    """
    Excel Document Handler (XLSX/XLS)
    
    Inherits from BaseHandler to manage config and image_processor at instance level.
    
    Usage:
        handler = ExcelHandler(config=config, image_processor=image_processor)
        text = handler.extract_text(current_file)
    """
    
    def extract_text(
        self,
        current_file: "CurrentFile",
        extract_metadata: bool = True,
        **kwargs
    ) -> str:
        """
        Extract text from Excel file.
        
        Args:
            current_file: CurrentFile dict containing file info and binary data
            extract_metadata: Whether to extract metadata
            **kwargs: Additional options
            
        Returns:
            Extracted text
        """
        file_path = current_file.get("file_path", "unknown")
        ext = current_file.get("file_extension", os.path.splitext(file_path)[1]).lower()
        # Normalize extension (remove leading dot if present)
        ext = ext.lstrip('.')
        self.logger.info(f"Excel processing: {file_path}, ext: {ext}")

        if ext == 'xlsx':
            return self._extract_xlsx(current_file, extract_metadata)
        elif ext == 'xls':
            return self._extract_xls(current_file, extract_metadata)
        else:
            raise ValueError(f"Unsupported Excel format: {ext}")
    
    def _extract_xlsx(
        self,
        current_file: "CurrentFile",
        extract_metadata: bool = True
    ) -> str:
        """XLSX file processing."""
        file_path = current_file.get("file_path", "unknown")
        self.logger.info(f"XLSX processing: {file_path}")

        try:
            # Open from stream to avoid path encoding issues
            file_stream = self.get_file_stream(current_file)
            wb = load_workbook(file_stream, data_only=True)
            preload = self._preload_xlsx_data(current_file, wb, extract_metadata)

            result_parts = [preload["metadata_str"]] if preload["metadata_str"] else []
            processed_images: Set[str] = set()
            stats = {"charts": 0, "images": 0, "textboxes": 0}

            for sheet_name in wb.sheetnames:
                sheet_result = self._process_xlsx_sheet(
                    wb[sheet_name], sheet_name, preload, processed_images, stats
                )
                result_parts.append(sheet_result)

            remaining = self._process_remaining_charts(
                preload["charts"], preload["chart_idx"], processed_images, stats
            )
            if remaining:
                result_parts.append(remaining)

            result = "".join(result_parts)
            self.logger.info(
                f"XLSX processing completed: {len(wb.sheetnames)} sheets, "
                f"{stats['charts']} charts, {stats['images']} images"
            )
            return result

        except Exception as e:
            self.logger.error(f"Error in XLSX processing: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            raise

    def _extract_xls(
        self,
        current_file: "CurrentFile",
        extract_metadata: bool = True
    ) -> str:
        """XLS file processing."""
        file_path = current_file.get("file_path", "unknown")
        self.logger.info(f"XLS processing: {file_path}")

        try:
            # xlrd can open from file_contents (bytes)
            file_data = current_file.get("file_data", b"")
            wb = xlrd.open_workbook(file_contents=file_data, formatting_info=True)
            result_parts = []

            if extract_metadata:
                metadata = extract_xls_metadata(wb)
                metadata_str = format_metadata(metadata)
                if metadata_str:
                    result_parts.append(metadata_str + "\n\n")

            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                sheet_tag = self.create_sheet_tag(ws.name)
                result_parts.append(f"\n{sheet_tag}\n")

                table_contents = convert_xls_objects_to_tables(ws, wb)
                if table_contents:
                    for i, table_content in enumerate(table_contents, 1):
                        if len(table_contents) > 1:
                            result_parts.append(f"\n[Table {i}]\n{table_content}\n")
                        else:
                            result_parts.append(f"\n{table_content}\n")

            result = "".join(result_parts)
            self.logger.info(f"XLS processing completed: {wb.nsheets} sheets")
            return result

        except Exception as e:
            self.logger.error(f"Error in XLS processing: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            raise

    def _preload_xlsx_data(
        self, current_file: "CurrentFile", wb, extract_metadata: bool
    ) -> Dict[str, Any]:
        """Extract preprocessing data from XLSX file."""
        file_path = current_file.get("file_path", "unknown")
        result = {
            "metadata_str": "",
            "charts": [],
            "images_data": [],
            "textboxes_by_sheet": {},
            "chart_idx": 0,
        }

        if extract_metadata:
            metadata = extract_xlsx_metadata(wb)
            result["metadata_str"] = format_metadata(metadata)
            if result["metadata_str"]:
                result["metadata_str"] += "\n\n"

        # NOTE: These helper functions still require file_path for now
        # They will need to be updated to use BytesIO streams in future
        result["charts"] = extract_charts_from_xlsx(file_path)
        result["images_data"] = extract_images_from_xlsx(file_path)
        result["textboxes_by_sheet"] = extract_textboxes_from_xlsx(file_path)

        return result

    def _process_xlsx_sheet(
        self, ws, sheet_name: str, preload: Dict[str, Any],
        processed_images: Set[str], stats: Dict[str, int]
    ) -> str:
        """Process a single XLSX sheet."""
        sheet_tag = self.create_sheet_tag(sheet_name)
        parts = [f"\n{sheet_tag}\n"]

        table_contents = convert_xlsx_objects_to_tables(ws)
        if table_contents:
            for i, table_content in enumerate(table_contents, 1):
                if len(table_contents) > 1:
                    parts.append(f"\n[Table {i}]\n{table_content}\n")
                else:
                    parts.append(f"\n{table_content}\n")

        # Chart processing
        if hasattr(ws, '_charts') and ws._charts:
            charts = preload["charts"]
            for chart in ws._charts:
                if preload["chart_idx"] < len(charts):
                    chart_data = charts[preload["chart_idx"]]
                    chart_output = process_chart(
                        chart_data,
                        processed_images,
                        self.image_processor.save_image
                    )
                    if chart_output:
                        parts.append(f"\n{chart_output}\n")
                        stats["charts"] += 1
                    preload["chart_idx"] += 1

        # Image processing
        sheet_images = get_sheet_images(ws, preload["images_data"], "")
        for image_data, anchor in sheet_images:
            if image_data:
                image_tag = self.image_processor.save_image(image_data)
                if image_tag:
                    parts.append(f"\n{image_tag}\n")
                    stats["images"] += 1

        # Textbox processing
        textboxes = preload["textboxes_by_sheet"].get(sheet_name, [])
        for tb in textboxes:
            if tb.get("text"):
                parts.append(f"\n[Textbox] {tb['text']}\n")
                stats["textboxes"] += 1

        return "".join(parts)

    def _process_remaining_charts(
        self, charts: List, chart_idx: int,
        processed_images: Set[str], stats: Dict[str, int]
    ) -> str:
        """Process remaining charts."""
        parts = []
        while chart_idx < len(charts):
            chart_data = charts[chart_idx]
            chart_output = process_chart(
                chart_data,
                processed_images,
                self.image_processor.save_image
            )
            if chart_output:
                parts.append(f"\n{chart_output}\n")
                stats["charts"] += 1
            chart_idx += 1
        return "".join(parts)


__all__ = ["ExcelHandler"]
